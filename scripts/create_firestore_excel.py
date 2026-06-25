#!/usr/bin/env python3
"""
create_firestore_excel.py
Membuat Excel dengan Power Query M code untuk semua koleksi Firestore sj-monitor.
Jalankan: python create_firestore_excel.py
Output: C:\Project\sj-monitor-firestore-query.xlsx
"""

import io, zipfile, uuid, re, os
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# COLORS & STYLES
# ─────────────────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E5EA6"
C_ACCENT      = "4472C4"
C_LIGHT_BLUE  = "D9E1F2"
C_HEADER_BG   = "2E5EA6"
C_PARAM_BG    = "EBF0FA"
C_INPUT_BG    = "FFFDE7"   # kuning muda  = input user
C_LOCK_BG     = "F2F2F2"   # abu2 = readonly
C_GREEN_BG    = "E2EFDA"
C_WHITE       = "FFFFFF"
C_GREEN_TXT   = "375623"
C_RED_TXT     = "9C0006"
C_ORANGE      = "C65911"

def cell_style(cell, bold=False, font_color=None, bg=None,
               h_align="left", v_align="center", wrap=False, size=10, font_name="Arial"):
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(bold=bold, color=(font_color or "000000"), name=font_name, size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)

def header_row(ws, row, cols_vals, bg=C_HEADER_BG, font_color=C_WHITE):
    for col, val in enumerate(cols_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        cell_style(c, bold=True, font_color=font_color, bg=bg, h_align="center")
    thin = Side(style="thin", color="FFFFFF")
    for col in range(1, len(cols_vals) + 1):
        ws.cell(row=row, column=col).border = Border(
            bottom=thin, top=thin, left=thin, right=thin
        )

def thin_border_range(ws, min_row, max_row, min_col, max_col, color="AAAAAA"):
    s = Side(style="thin", color=color)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# M CODE  (Power Query)
# ─────────────────────────────────────────────────────────────────────────────
QUERIES = [
    "fn_GetParam", "fn_FirebaseToken", "fn_ParseFirestoreValue",
    "fn_FirestoreQuery", "fn_FirestoreGetAll",
    "tbl_Supir", "tbl_Trucks", "tbl_Rute", "tbl_Material", "tbl_TarifRute",
    "tbl_Users", "tbl_SuratJalan", "tbl_Biaya", "tbl_Invoice",
    "tbl_UangMuka", "tbl_Transaksi", "tbl_HistoryLog",
]
TABLE_QUERIES = [q for q in QUERIES if q.startswith("tbl_")]

M_SECTION = r'''section Section1;

// ── Baca satu nilai dari tabel Config ──────────────────────────────────────
shared fn_GetParam = (paramName as text) =>
let
    tbl = Excel.CurrentWorkbook(){[Name="tbl_Config"]}[Content],
    row = Table.SelectRows(tbl, each [Parameter] = paramName),
    val = if Table.RowCount(row) > 0 then row{0}[Nilai] else null
in
    val;

// ── Ambil Firebase ID Token via email/password (masa berlaku 1 jam) ────────
shared fn_FirebaseToken =
let
    api_key  = fn_GetParam("api_key"),
    email    = fn_GetParam("email"),
    password = fn_GetParam("password"),
    body     = Text.ToBinary(
        "{""email"":""" & email & """,""password"":""" & password
        & """,""returnSecureToken"":true}",
        TextEncoding.Utf8
    ),
    response = Web.Contents(
        "https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key=" & api_key,
        [ Headers = [#"Content-Type" = "application/json"], Content = body ]
    )
in
    Json.Document(response)[idToken];

// ── Parse satu field value dari Firestore ─────────────────────────────────
shared fn_ParseFirestoreValue = (v as record) =>
    if      Record.HasFields(v, "stringValue")    then v[stringValue]
    else if Record.HasFields(v, "integerValue")   then Number.FromText(Text.From(v[integerValue]))
    else if Record.HasFields(v, "doubleValue")    then v[doubleValue]
    else if Record.HasFields(v, "booleanValue")   then v[booleanValue]
    else if Record.HasFields(v, "timestampValue") then v[timestampValue]
    else null;

// ── Query Firestore dengan filter tanggal (date range) ────────────────────
// collectionId : nama koleksi Firestore
// dateField    : nama field tanggal (e.g. "tanggalSJ", "tanggal", "tglInvoice")
// startDate    : "YYYY-MM-DD"
// endDate      : "YYYY-MM-DD"
// queryLimit   : opsional, default 10000
shared fn_FirestoreQuery =
    (collectionId as text, dateField as text, startDate as text, endDate as text,
     optional queryLimit as number) =>
let
    lim        = if queryLimit = null then 10000 else queryLimit,
    project_id = fn_GetParam("project_id"),
    token      = fn_FirebaseToken,
    filters    = {
        [ fieldFilter = [ field = [fieldPath = dateField],
                          op    = "GREATER_THAN_OR_EQUAL",
                          value = [stringValue = startDate] ] ],
        [ fieldFilter = [ field = [fieldPath = dateField],
                          op    = "LESS_THAN_OR_EQUAL",
                          value = [stringValue = endDate] ] ]
    },
    queryBody  = [
        structuredQuery = [
            from    = {[collectionId = collectionId]},
            where   = [compositeFilter = [op = "AND", filters = filters]],
            orderBy = {[field = [fieldPath = dateField], direction = "DESCENDING"]},
            limit   = lim
        ]
    ],
    url        = "https://firestore.googleapis.com/v1/projects/" & project_id
                 & "/databases/(default)/documents:runQuery",
    response   = Web.Contents(url, [
        Headers = [ #"Content-Type" = "application/json",
                    Authorization   = "Bearer " & token ],
        Content = Json.FromValue(queryBody)
    ]),
    docs       = Json.Document(response),
    parsed     = List.Transform(docs, (item) =>
        if not Record.HasFields(item, "document") then null
        else let
            doc   = item[document],
            flds  = if Record.HasFields(doc, "fields") then doc[fields] else [],
            names = Record.FieldNames(flds),
            vals  = List.Transform(names, each fn_ParseFirestoreValue(Record.Field(flds, _))),
            rec   = Record.FromList(vals, names),
            docId = List.Last(Text.Split(doc[name], "/"))
        in Record.AddField(rec, "_docId", docId)
    ),
    nonNull    = List.Select(parsed, each _ <> null),
    result     = if List.Count(nonNull) > 0 then Table.FromRecords(nonNull)
                 else #table({},{})
in
    result;

// ── Ambil SEMUA dokumen dari satu koleksi (tanpa filter tanggal) ──────────
// Digunakan untuk master data: supir, trucks, rute, material, tarif_rute, users
shared fn_FirestoreGetAll = (collectionId as text) =>
let
    project_id = fn_GetParam("project_id"),
    token      = fn_FirebaseToken,
    queryBody  = [ structuredQuery = [ from = {[collectionId = collectionId]}, limit = 1000 ] ],
    url        = "https://firestore.googleapis.com/v1/projects/" & project_id
                 & "/databases/(default)/documents:runQuery",
    response   = Web.Contents(url, [
        Headers = [ #"Content-Type" = "application/json",
                    Authorization   = "Bearer " & token ],
        Content = Json.FromValue(queryBody)
    ]),
    docs       = Json.Document(response),
    parsed     = List.Transform(docs, (item) =>
        if not Record.HasFields(item, "document") then null
        else let
            doc   = item[document],
            flds  = if Record.HasFields(doc, "fields") then doc[fields] else [],
            names = Record.FieldNames(flds),
            vals  = List.Transform(names, each fn_ParseFirestoreValue(Record.Field(flds, _))),
            rec   = Record.FromList(vals, names),
            docId = List.Last(Text.Split(doc[name], "/"))
        in Record.AddField(rec, "_docId", docId)
    ),
    nonNull    = List.Select(parsed, each _ <> null),
    result     = if List.Count(nonNull) > 0 then Table.FromRecords(nonNull)
                 else #table({},{})
in
    result;

// ════════════════════════════════════════════════════════════════════════════
// MASTER DATA  (tidak difilter tanggal/supir/truck/rute)
// ════════════════════════════════════════════════════════════════════════════

shared tbl_Supir =
let
    raw    = fn_FirestoreGetAll("supir"),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

shared tbl_Trucks =
let
    raw    = fn_FirestoreGetAll("trucks"),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

shared tbl_Rute =
let
    raw    = fn_FirestoreGetAll("rute"),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

shared tbl_Material =
let
    raw    = fn_FirestoreGetAll("material"),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

shared tbl_TarifRute =
let
    raw    = fn_FirestoreGetAll("tarif_rute"),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

shared tbl_Users =
let
    raw    = fn_FirestoreGetAll("users"),
    active = if Table.HasColumns(raw, {"deletedAt"})
             then Table.SelectRows(raw, each ([deletedAt] = null or [deletedAt] = ""))
             else raw
in active;

// ════════════════════════════════════════════════════════════════════════════
// SURAT JALAN  — filter: tanggal + supir + truck (noPol) + rute
// Filter opsional: kosongkan kolom di Config untuk mengambil semua data
// ════════════════════════════════════════════════════════════════════════════
shared tbl_SuratJalan =
let
    start_date   = fn_GetParam("start_date"),
    end_date     = fn_GetParam("end_date"),
    f_supir      = fn_GetParam("filter_supir"),
    f_truck      = fn_GetParam("filter_truck"),
    f_rute       = fn_GetParam("filter_rute"),

    // 1. Ambil data dari Firestore (filter tanggal server-side)
    raw          = fn_FirestoreQuery("surat_jalan", "tanggalSJ", start_date, end_date),

    // 2. Buang soft-deleted
    actv         = if Table.HasColumns(raw, {"isActive"})
                   then Table.SelectRows(raw, each [isActive] <> false)
                   else raw,

    // 3. Filter supir (namaSupir sudah denormalized di dokumen SJ)
    f1           = if f_supir <> null and f_supir <> ""
                   then if Table.HasColumns(actv, {"namaSupir"})
                        then Table.SelectRows(actv, each
                             Text.Contains(
                                 Text.Lower(if [namaSupir] = null then "" else [namaSupir]),
                                 Text.Lower(f_supir)
                             ))
                        else actv
                   else actv,

    // 4. Join dengan tbl_Trucks untuk dapat noPol, lalu filter
    withTruck    = if Table.HasColumns(f1, {"truckId"})
                      and Table.HasColumns(tbl_Trucks, {"id", "noPol"})
                   then
                       let jn = Table.NestedJoin(f1, {"truckId"}, tbl_Trucks, {"id"},
                                                 "_trk", JoinKind.LeftOuter)
                       in Table.ExpandTableColumn(jn, "_trk", {"noPol"}, {"noPol"})
                   else Table.AddColumn(f1, "noPol", each null, type text),

    f2           = if f_truck <> null and f_truck <> ""
                   then Table.SelectRows(withTruck, each
                        Text.Contains(
                            Text.Lower(if [noPol] = null then "" else [noPol]),
                            Text.Lower(f_truck)
                        ))
                   else withTruck,

    // 5. Join dengan tbl_Rute untuk dapat nama rute, lalu filter
    withRute     = if Table.HasColumns(f2, {"ruteId"})
                      and Table.HasColumns(tbl_Rute, {"id", "rute"})
                   then
                       let jn = Table.NestedJoin(f2, {"ruteId"}, tbl_Rute, {"id"},
                                                 "_rte", JoinKind.LeftOuter)
                       in Table.ExpandTableColumn(jn, "_rte", {"rute"}, {"namaRute"})
                   else Table.AddColumn(f2, "namaRute", each null, type text),

    f3           = if f_rute <> null and f_rute <> ""
                   then Table.SelectRows(withRute, each
                        Text.Contains(
                            Text.Lower(if [namaRute] = null then "" else [namaRute]),
                            Text.Lower(f_rute)
                        ))
                   else withRute
in f3;

// ════════════════════════════════════════════════════════════════════════════
// BIAYA  — filter: tanggal saja
// ════════════════════════════════════════════════════════════════════════════
shared tbl_Biaya =
let
    raw    = fn_FirestoreQuery("biaya", "tanggal",
                               fn_GetParam("start_date"), fn_GetParam("end_date")),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

// ════════════════════════════════════════════════════════════════════════════
// INVOICE  — gabung koleksi "invoice" + "invoices" (legacy), deduplikasi
// ════════════════════════════════════════════════════════════════════════════
shared tbl_Invoice =
let
    start  = fn_GetParam("start_date"),
    end    = fn_GetParam("end_date"),
    r1     = fn_FirestoreQuery("invoice",  "tglInvoice", start, end),
    r2     = fn_FirestoreQuery("invoices", "tglInvoice", start, end),
    comb   = Table.Combine({r1, r2}),
    actv   = if Table.HasColumns(comb, {"isActive"})
             then Table.SelectRows(comb, each [isActive] <> false)
             else comb,
    sorted = if Table.HasColumns(actv, {"updatedAt"})
             then Table.Sort(actv, {{"updatedAt", Order.Descending}})
             else actv,
    dedup  = if Table.HasColumns(sorted, {"noInvoice"})
             then Table.Distinct(sorted, {"noInvoice"})
             else sorted
in dedup;

// ════════════════════════════════════════════════════════════════════════════
// UANG MUKA  — filter: tanggal saja
// ════════════════════════════════════════════════════════════════════════════
shared tbl_UangMuka =
let
    raw    = fn_FirestoreQuery("uang_muka", "tanggal",
                               fn_GetParam("start_date"), fn_GetParam("end_date")),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

// ════════════════════════════════════════════════════════════════════════════
// TRANSAKSI  — filter: tanggal saja
// ════════════════════════════════════════════════════════════════════════════
shared tbl_Transaksi =
let
    raw    = fn_FirestoreQuery("transaksi", "tanggal",
                               fn_GetParam("start_date"), fn_GetParam("end_date")),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;

// ════════════════════════════════════════════════════════════════════════════
// HISTORY LOG  — filter: tanggal, limit 1000 baris terbaru
// ════════════════════════════════════════════════════════════════════════════
shared tbl_HistoryLog =
let
    raw    = fn_FirestoreQuery("history_log", "timestamp",
                               fn_GetParam("start_date"), fn_GetParam("end_date"), 1000),
    active = if Table.HasColumns(raw, {"isActive"})
             then Table.SelectRows(raw, each [isActive] <> false)
             else raw
in active;
'''

# ─────────────────────────────────────────────────────────────────────────────
# BUILD DATAMASHUP BINARY (inner ZIP)
# ─────────────────────────────────────────────────────────────────────────────
def _build_package_xml():
    entries = "\n".join(
        f'    <Query Name="{q}"><Description/>'
        f'<IsParameterQuery>false</IsParameterQuery>'
        f'<IsDirectQuery>false</IsDirectQuery></Query>'
        for q in QUERIES
    )
    return f"""<?xml version="1.0" encoding="utf-8"?>
<Package xmlns:xsd="http://www.w3.org/2001/XMLSchema"
         xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
         xmlns="http://schemas.microsoft.com/DataMashup/2013">
  <Version Major="2" Minor="21" Build="0" Revision="0"/>
  <MinVersion Major="2" Minor="21" Build="0" Revision="0"/>
  <Culture>en-US</Culture>
  <SafeCombine>false</SafeCombine>
  <QueryGroups/>
  <Queries>
{entries}
  </Queries>
</Package>"""

_DM_CONTENT_TYPES = """<?xml version="1.0" encoding="utf-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/Formulas/Section1.m"
            ContentType="application/vnd.ms-excel.query+m"/>
  <Override PartName="/Config/Package.xml"
            ContentType="application/vnd.ms-excel.query-package+xml"/>
</Types>"""

_DM_RELS = """<?xml version="1.0" encoding="utf-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Type="http://schemas.microsoft.com/DataMashup/2013/Package"
                Target="Config/Package.xml" Id="R1"/>
  <Relationship Type="http://schemas.microsoft.com/DataMashup/2013/Section"
                Target="Formulas/Section1.m" Id="R2"/>
</Relationships>"""

def build_datamashup() -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", _DM_CONTENT_TYPES)
        zf.writestr("_rels/.rels", _DM_RELS)
        zf.writestr("Config/Package.xml", _build_package_xml())
        zf.writestr("Formulas/Section1.m", M_SECTION.encode("utf-8"))
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# BUILD connections.xml
# ─────────────────────────────────────────────────────────────────────────────
def build_connections_xml() -> str:
    conns = []
    for idx, name in enumerate(TABLE_QUERIES, 1):
        guid = "{" + str(uuid.uuid4()).upper() + "}"
        conns.append(f"""  <connection id="{idx}" name="Query - {name}"
              description="Power Query: {name}" type="5"
              refreshedVersion="3" background="1" savePassword="0">
    <dbPr connection="Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;Location={name};Extended Properties=&quot;&quot;"
          command="SELECT * FROM [{name}]"/>
    <extLst>
      <ext xmlns:x15="http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
           uri="{{DE250136-89BD-433C-8126-D09CA5730AF9}}">
        <x15:connection id="{guid}" type="powerQueryTable" dbCommandType="2"/>
      </ext>
    </extLst>
  </connection>""")
    body = "\n".join(conns)
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
             xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
             xmlns:x15="http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
             xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2">
{body}
</connections>"""

# ─────────────────────────────────────────────────────────────────────────────
# INJECT POWER QUERY into .xlsx (zip manipulation)
# ─────────────────────────────────────────────────────────────────────────────
def inject_power_query(xlsx_path: str):
    dm_bytes = build_datamashup()
    conn_xml = build_connections_xml().encode("utf-8")

    with zipfile.ZipFile(xlsx_path, "r") as src:
        parts = {n: src.read(n) for n in src.namelist()}

    # ── [Content_Types].xml ──────────────────────────────────────────────────
    ct = parts["[Content_Types].xml"].decode("utf-8")
    if "customXml/item1.xml" not in ct:
        ct = ct.replace(
            "</Types>",
            '  <Override PartName="/xl/customXml/item1.xml"'
            ' ContentType="application/octet-stream"/>\n</Types>'
        )
    if "connections.xml" not in ct:
        ct = ct.replace(
            "</Types>",
            '  <Override PartName="/xl/connections.xml" ContentType="application/'
            'vnd.openxmlformats-officedocument.spreadsheetml.connections+xml"/>\n</Types>'
        )
    parts["[Content_Types].xml"] = ct.encode("utf-8")

    # ── xl/_rels/workbook.xml.rels ───────────────────────────────────────────
    wr_key = "xl/_rels/workbook.xml.rels"
    wr = parts.get(wr_key, b"").decode("utf-8")
    existing_ids = re.findall(r'Id="rId(\d+)"', wr)
    next_id = max((int(x) for x in existing_ids), default=0) + 1

    extra_rels = ""
    if "DataMashup" not in wr:
        extra_rels += (
            f'  <Relationship Id="rId{next_id}" '
            f'Type="http://schemas.microsoft.com/DataMashup" '
            f'Target="customXml/item1.xml"/>\n'
        )
        next_id += 1
    if "connections" not in wr:
        extra_rels += (
            f'  <Relationship Id="rId{next_id}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            f'relationships/connections" Target="connections.xml"/>\n'
        )
    if extra_rels:
        wr = wr.replace("</Relationships>", extra_rels + "</Relationships>")
    parts[wr_key] = wr.encode("utf-8")

    # ── DataMashup binary & connections ──────────────────────────────────────
    parts["xl/customXml/item1.xml"] = dm_bytes
    parts["xl/connections.xml"] = conn_xml

    # ── Tulis ulang xlsx ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for name, data in parts.items():
            dst.writestr(name, data)
    with open(xlsx_path, "wb") as f:
        f.write(buf.getvalue())

# ─────────────────────────────────────────────────────────────────────────────
# BUILD EXCEL WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
def build_workbook(out_path: str):
    wb = Workbook()

    # ── 1. Sheet Panduan ──────────────────────────────────────────────────────
    ws_guide = wb.active
    ws_guide.title = "📋 Panduan"
    ws_guide.sheet_properties.tabColor = "4472C4"
    ws_guide.column_dimensions["A"].width = 3
    ws_guide.column_dimensions["B"].width = 28
    ws_guide.column_dimensions["C"].width = 55

    def guide_row(r, label, val="", lb="", lv=""):
        c = ws_guide.cell(r, 2, label)
        cell_style(c, bold=bool(lb), font_color=lb or "000000", bg=C_LIGHT_BLUE if lb else None,
                   h_align="left", size=10)
        c2 = ws_guide.cell(r, 3, val)
        cell_style(c2, font_color=lv or "000000", size=10, wrap=True)

    ws_guide.merge_cells("B1:C1")
    t = ws_guide.cell(1, 2, "📊 sj-monitor — Firestore Power Query Template")
    cell_style(t, bold=True, font_color=C_WHITE, bg=C_DARK_BLUE, size=14, h_align="left")
    ws_guide.row_dimensions[1].height = 30

    steps = [
        ("", ""),
        ("LANGKAH SETUP (lakukan sekali)", ""),
        ("1. Isi sheet ⚙️ Config", "Masukkan project_id, api_key, email, dan password Firebase Anda di kolom Nilai"),
        ("2. Atur rentang tanggal", "start_date dan end_date menentukan data yang ditarik (format YYYY-MM-DD)"),
        ("3. Opsional: isi filter", "filter_supir / filter_truck / filter_rute → isi nama/noPol untuk filter, kosongkan untuk semua data"),
        ("", ""),
        ("CARA REFRESH DATA", ""),
        ("4. Buka panel Query", "Data → Queries & Connections (atau Ctrl+Alt+F5 untuk refresh all)"),
        ("5. Klik kanan query", "Klik kanan nama query → Load To... → pilih sheet tujuan"),
        ("6. Refresh ulang", "Setelah mengubah Config, tekan Ctrl+Alt+F5 atau klik kanan → Refresh"),
        ("", ""),
        ("DAFTAR QUERIES / TABEL", ""),
        ("tbl_SuratJalan", "Surat Jalan — filter tanggal+supir+truck+rute"),
        ("tbl_Biaya", "Biaya — filter tanggal"),
        ("tbl_Invoice", "Invoice (merge 'invoice'+'invoices') — filter tanggal"),
        ("tbl_UangMuka", "Uang Muka — filter tanggal"),
        ("tbl_Transaksi", "Transaksi — filter tanggal"),
        ("tbl_HistoryLog", "History Log — filter tanggal, max 1000 baris"),
        ("tbl_Supir", "Master Supir — semua data aktif"),
        ("tbl_Trucks", "Master Truck — semua data aktif"),
        ("tbl_Rute", "Master Rute — semua data aktif"),
        ("tbl_Material", "Master Material — semua data aktif"),
        ("tbl_TarifRute", "Tarif Rute — semua data aktif"),
        ("tbl_Users", "Daftar User — tanpa soft-deleted"),
        ("", ""),
        ("CATATAN PENTING", ""),
        ("Token Firebase", "ID Token berlaku 1 jam. Jika refresh gagal, ubah salah satu nilai di Config lalu Refresh ulang"),
        ("Kuota Firestore", "Setiap Refresh = N read Firestore. Gunakan rentang tanggal sekecil mungkin"),
        ("Filter case-insensitive", "Filter supir/truck/rute menggunakan pencarian 'contains', bukan exact match"),
        ("Master data", "tbl_Supir, tbl_Trucks, tbl_Rute, tbl_Material tidak dibatasi tanggal"),
    ]

    for i, (lbl, val) in enumerate(steps, 2):
        is_section = val == "" and lbl != ""
        is_blank = lbl == ""
        bg = C_LIGHT_BLUE if is_section else None
        bold = is_section or lbl.startswith(("LANGKAH", "CARA", "DAFTAR", "CATATAN"))
        fc = C_DARK_BLUE if is_section else ("374151" if lbl.startswith("tbl_") else "000000")
        guide_row(i, lbl, val, lb=(C_DARK_BLUE if is_section else ""), lv="")
        c = ws_guide.cell(i, 2)
        cell_style(c, bold=bold, font_color=fc,
                   bg=C_LIGHT_BLUE if is_section else (C_PARAM_BG if lbl.startswith("tbl_") else None),
                   size=10)
        ws_guide.row_dimensions[i].height = 18 if not is_blank else 8

    # ── 2. Sheet Config ───────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("⚙️ Config")
    ws_cfg.sheet_properties.tabColor = "ED7D31"
    ws_cfg.column_dimensions["A"].width = 3
    ws_cfg.column_dimensions["B"].width = 22
    ws_cfg.column_dimensions["C"].width = 38
    ws_cfg.column_dimensions["D"].width = 45
    ws_cfg.column_dimensions["E"].width = 3

    # Title
    ws_cfg.merge_cells("B1:D1")
    t = ws_cfg.cell(1, 2, "⚙️  Konfigurasi — Ubah nilai di kolom Nilai, lalu Refresh")
    cell_style(t, bold=True, font_color=C_WHITE, bg=C_ORANGE, size=12, h_align="left")
    ws_cfg.row_dimensions[1].height = 28

    # Header tabel
    header_row(ws_cfg, 2, ["", "Parameter", "Nilai", "Keterangan"], bg=C_DARK_BLUE)
    ws_cfg.cell(2, 1).fill = PatternFill("solid", start_color=C_DARK_BLUE)

    today = date.today()
    params = [
        # (Parameter, Nilai_default, Keterangan, is_input)
        ("─── Kredensial Firebase ───", "", "", False),
        ("project_id",  "",  "Firebase Project ID — lihat di Firebase Console → Project Settings", True),
        ("api_key",     "",  "Firebase Web API Key — dari Project Settings → General → Web API Key", True),
        ("email",       "",  "Email akun Firebase yang punya akses baca Firestore", True),
        ("password",    "",  "Password akun tersebut (disimpan plain text — gunakan akun khusus/read-only)", True),
        ("─── Filter Tanggal ───", "", "", False),
        ("start_date",  f"{today.year}-01-01",  "Tanggal mulai data — format YYYY-MM-DD", True),
        ("end_date",    today.strftime("%Y-%m-%d"), "Tanggal akhir data — format YYYY-MM-DD", True),
        ("─── Filter Opsional ───", "", "", False),
        ("filter_supir","",  "Nama supir (sebagian sudah cukup, misal: 'Budi') — kosongkan untuk semua", True),
        ("filter_truck","",  "No. polisi truck (sebagian, misal: 'B 1234') — kosongkan untuk semua", True),
        ("filter_rute", "",  "Nama rute (sebagian, misal: 'Cibitung') — kosongkan untuk semua", True),
    ]

    tbl_start = 3
    for i, (param, val, ket, is_inp) in enumerate(params):
        row = tbl_start + i
        is_section = param.startswith("─")
        if is_section:
            ws_cfg.merge_cells(f"B{row}:D{row}")
            c = ws_cfg.cell(row, 2, param)
            cell_style(c, bold=True, font_color=C_WHITE, bg=C_ACCENT, h_align="left", size=10)
            ws_cfg.row_dimensions[row].height = 20
            continue

        ws_cfg.cell(row, 1).fill = PatternFill("solid", start_color=C_PARAM_BG)
        c_param = ws_cfg.cell(row, 2, param)
        cell_style(c_param, bold=True, font_color="1F3864", bg=C_PARAM_BG, size=10)

        c_val = ws_cfg.cell(row, 3, val)
        cell_style(c_val, bg=C_INPUT_BG, size=10, h_align="left")
        # Tandai sebagai input (warna biru per konvensi financial model)
        c_val.font = Font(color="000099", name="Arial", size=10, bold=False)

        c_ket = ws_cfg.cell(row, 4, ket)
        cell_style(c_ket, font_color="555555", bg=C_LOCK_BG, size=9, wrap=True)
        ws_cfg.row_dimensions[row].height = 22

    # Named table tbl_Config
    tbl_end = tbl_start + len(params) - 1
    # Hitung baris tabel yang bukan section header
    data_rows = [p for p in params if not p[0].startswith("─")]
    # Kita perlu buat named table Excel dari range B{param_start}:C{param_end}
    # tapi named table butuh baris yang konsisten. Buat tabel terpisah di bawah.
    # Row untuk named table: buat block tersendiri
    tbl_named_row_start = tbl_end + 3
    ws_cfg.cell(tbl_named_row_start, 2, "Parameter").font = Font(bold=True, name="Arial", size=9, color=C_DARK_BLUE)
    ws_cfg.cell(tbl_named_row_start, 3, "Nilai").font = Font(bold=True, name="Arial", size=9, color=C_DARK_BLUE)
    ws_cfg.cell(tbl_named_row_start-1, 2,
                "↓  Tabel tersembunyi di bawah ini digunakan Power Query (jangan hapus)").font = \
        Font(italic=True, color="888888", name="Arial", size=8)

    named_data_rows = [(p, v) for p, v, *_ in params if not p.startswith("─")]
    for j, (p, v) in enumerate(named_data_rows):
        r = tbl_named_row_start + 1 + j
        ws_cfg.cell(r, 2, p)
        ws_cfg.cell(r, 3, v)
        ws_cfg.row_dimensions[r].height = 16

    named_end = tbl_named_row_start + len(named_data_rows)

    from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
    tbl = XlTable(
        displayName="tbl_Config",
        ref=f"B{tbl_named_row_start}:C{named_end}",
    )
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws_cfg.add_table(tbl)

    # Hide the helper rows (optional - keep visible for transparency)
    # for r in range(tbl_named_row_start-1, named_end+1):
    #     ws_cfg.row_dimensions[r].hidden = True

    # ── 3. Sheet M Code (per query, individual paste) ────────────────────────
    ws_mcode = wb.create_sheet("📝 M Code")
    ws_mcode.sheet_properties.tabColor = "70AD47"
    ws_mcode.column_dimensions["A"].width = 3
    ws_mcode.column_dimensions["B"].width = 26   # Nama query / label
    ws_mcode.column_dimensions["C"].width = 100  # M code

    # ── Title
    ws_mcode.merge_cells("B1:C1")
    t = ws_mcode.cell(1, 2, "📝  M Code — Buat setiap query SATU PER SATU di Power Query Editor")
    cell_style(t, bold=True, font_color=C_WHITE, bg=C_GREEN_TXT, size=12, h_align="left")
    ws_mcode.row_dimensions[1].height = 28

    # ── Instruksi singkat
    instrs = [
        "⚠  JANGAN paste seluruh halaman ini ke satu query — error 'Invalid expression'!",
        "CARA YANG BENAR:",
        "  1. Data → Get Data → Launch Power Query Editor",
        "  2. Home → New Source → Other Sources → Blank Query",
        "  3. Klik kanan query baru di panel kiri → Rename → ketik Nama Query (kolom kuning)",
        "  4. Klik kanan query → Advanced Editor → HAPUS semua isi → paste M Code (kolom hijau)",
        "  5. Klik Done. Ulangi untuk setiap query sesuai urutan di bawah.",
        "  6. Close & Load saat semua query sudah dibuat.",
        "  CATATAN: Buat query SESUAI URUTAN (fn_* dulu, lalu tbl_*) karena ada dependensi.",
    ]
    for i, txt in enumerate(instrs, 2):
        ws_mcode.merge_cells(f"B{i}:C{i}")
        c = ws_mcode.cell(i, 2, txt)
        is_warn = txt.startswith("⚠")
        is_note = txt.startswith("  CATATAN")
        is_head = txt == "CARA YANG BENAR:"
        bg  = "FFF2CC" if is_warn else (C_LIGHT_BLUE if is_head else None)
        fc  = "9C0006" if is_warn else (C_DARK_BLUE if is_head else ("8B4513" if is_note else "374151"))
        bld = is_warn or is_head or is_note
        cell_style(c, bold=bld, font_color=fc, bg=bg, size=9 if not is_warn else 10, h_align="left")
        ws_mcode.row_dimensions[i].height = 16
    ws_mcode.row_dimensions[2].height = 20  # warning baris lebih tinggi

    # ── Header kolom tabel
    hdr_row = 2 + len(instrs) + 1
    ws_mcode.merge_cells(f"A{hdr_row}:A{hdr_row}")
    for col, (val, bg, fc) in enumerate([
        ("", C_DARK_BLUE, C_WHITE),
        ("Nama Query  (rename blank query menjadi ini ↓)", C_DARK_BLUE, C_WHITE),
        ("M Code — copy SEMUA baris kode ini → paste ke Advanced Editor", C_DARK_BLUE, C_WHITE),
    ], 1):
        c = ws_mcode.cell(hdr_row, col, val)
        cell_style(c, bold=True, font_color=fc, bg=bg, h_align="left", size=9)
    ws_mcode.row_dimensions[hdr_row].height = 20

    # ── Parse M_SECTION → individual (name, code) list
    def _parse_section(txt):
        out, cur_name, buf = [], None, []
        for line in txt.splitlines():
            m = re.match(r'^shared\s+(\w+)\s*=(.*)', line)
            if m:
                if cur_name is not None:
                    code = '\n'.join(buf).strip().rstrip(';').strip()
                    out.append((cur_name, code))
                cur_name = m.group(1)
                buf = [m.group(2)]  # rest of same line (may be empty)
            elif cur_name is not None:
                buf.append(line)
        if cur_name is not None:
            code = '\n'.join(buf).strip().rstrip(';').strip()
            out.append((cur_name, code))
        return out

    # Dependency info per query (untuk label)
    DEPS = {
        "fn_GetParam":           "—",
        "fn_FirebaseToken":      "fn_GetParam",
        "fn_ParseFirestoreValue":"—",
        "fn_FirestoreQuery":     "fn_GetParam, fn_FirebaseToken, fn_ParseFirestoreValue",
        "fn_FirestoreGetAll":    "fn_GetParam, fn_FirebaseToken, fn_ParseFirestoreValue",
        "tbl_Supir":             "fn_FirestoreGetAll",
        "tbl_Trucks":            "fn_FirestoreGetAll",
        "tbl_Rute":              "fn_FirestoreGetAll",
        "tbl_Material":          "fn_FirestoreGetAll",
        "tbl_TarifRute":         "fn_FirestoreGetAll",
        "tbl_Users":             "fn_FirestoreGetAll",
        "tbl_SuratJalan":        "fn_FirestoreQuery, tbl_Trucks, tbl_Rute",
        "tbl_Biaya":             "fn_FirestoreQuery",
        "tbl_Invoice":           "fn_FirestoreQuery",
        "tbl_UangMuka":          "fn_FirestoreQuery",
        "tbl_Transaksi":         "fn_FirestoreQuery",
        "tbl_HistoryLog":        "fn_FirestoreQuery",
    }
    FN_QUERIES = {"fn_GetParam","fn_FirebaseToken","fn_ParseFirestoreValue",
                  "fn_FirestoreQuery","fn_FirestoreGetAll"}

    queries = _parse_section(M_SECTION)
    cur_row = hdr_row + 1

    # Separator helper
    def _sep(ws, r, label, bg):
        ws.merge_cells(f"B{r}:C{r}")
        c = ws.cell(r, 2, label)
        cell_style(c, bold=True, font_color=C_WHITE, bg=bg, size=9)
        ws.cell(r, 1).fill = PatternFill("solid", start_color=bg)
        ws.row_dimensions[r].height = 16
        return r + 1

    cur_row = _sep(ws_mcode, cur_row, "BUAT TERLEBIH DAHULU: Helper Functions (fn_*)", C_ACCENT)

    fn_done = False
    for q_name, q_code in queries:
        is_fn = q_name in FN_QUERIES
        # Separator sebelum tbl_* pertama
        if not is_fn and not fn_done:
            fn_done = True
            cur_row = _sep(ws_mcode, cur_row, "BUAT SETELAH SEMUA fn_*: Table Queries (tbl_*)", C_GREEN_TXT)

        # Kolom A (gutter warna)
        gutter_bg = C_ACCENT if is_fn else C_GREEN_TXT
        ws_mcode.cell(cur_row, 1).fill = PatternFill("solid", start_color=gutter_bg)

        # Kolom B — nama query + dependensi
        dep_text = DEPS.get(q_name, "")
        name_label = f"{q_name}\n\nDep: {dep_text}"
        c_name = ws_mcode.cell(cur_row, 2, name_label)
        name_bg = C_LIGHT_BLUE if is_fn else C_GREEN_BG
        name_fc = C_DARK_BLUE if is_fn else C_GREEN_TXT
        cell_style(c_name, bold=True, font_color=name_fc, bg=name_bg, size=9,
                   h_align="left", v_align="top", wrap=True)

        # Kolom C — M code, baris per baris dalam satu cell (wrap)
        c_code = ws_mcode.cell(cur_row, 3, q_code)
        code_bg = "EBF7FF" if is_fn else "EFFFEE"
        cell_style(c_code, bg=code_bg, size=9, h_align="left", v_align="top", wrap=True)
        c_code.font = Font(name="Courier New", size=9, color="1F3864")

        # Tinggi baris: estimasi berdasarkan jumlah baris kode
        n_lines = q_code.count('\n') + 1
        row_h = max(n_lines * 13, 40)
        ws_mcode.row_dimensions[cur_row].height = row_h

        # Border tipis di sekeliling
        s = Side(style="thin", color="CCCCCC")
        for col in range(1, 4):
            ws_mcode.cell(cur_row, col).border = Border(left=s, right=s, top=s, bottom=s)

        cur_row += 1  # empty gap
        ws_mcode.row_dimensions[cur_row].height = 6
        cur_row += 1

    # ── 4. Sheet per tabel (placeholder) ─────────────────────────────────────
    SHEET_DEFS = [
        ("🚚 Surat Jalan", "tbl_SuratJalan",  C_MID_BLUE,  "tanggalSJ + namaSupir + noPol + namaRute"),
        ("💰 Biaya",       "tbl_Biaya",        "7030A0",    "tanggal"),
        ("🧾 Invoice",     "tbl_Invoice",      "C00000",    "tglInvoice"),
        ("💵 Uang Muka",   "tbl_UangMuka",     "843C0C",    "tanggal"),
        ("💳 Transaksi",   "tbl_Transaksi",    "1F497D",    "tanggal"),
        ("📜 History Log", "tbl_HistoryLog",   "595959",    "timestamp (max 1000)"),
        ("👤 Supir",       "tbl_Supir",        C_GREEN_TXT, "—  (semua master data)"),
        ("🚛 Trucks",      "tbl_Trucks",       C_GREEN_TXT, "—  (semua master data)"),
        ("🗺️ Rute",        "tbl_Rute",         C_GREEN_TXT, "—  (semua master data)"),
        ("🪨 Material",    "tbl_Material",     C_GREEN_TXT, "—  (semua master data)"),
        ("📊 Tarif Rute",  "tbl_TarifRute",    C_GREEN_TXT, "—  (semua master data)"),
        ("👥 Users",       "tbl_Users",        "595959",    "—  (semua user aktif)"),
    ]

    for sheet_name, query_name, tab_color, filter_desc in SHEET_DEFS:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_color
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 60

        ws.merge_cells("A1:B1")
        t = ws.cell(1, 1, f"  {sheet_name}  ←  Query: {query_name}")
        cell_style(t, bold=True, font_color=C_WHITE, bg=tab_color, size=12, h_align="left")
        ws.row_dimensions[1].height = 28

        ws.cell(2, 1, "Filter aktif: " + filter_desc)
        ws.cell(2, 1).font = Font(italic=True, color="555555", name="Arial", size=9)

        ws.cell(4, 1, "⬇  Cara load data ke sheet ini:")
        ws.cell(4, 1).font = Font(bold=True, color=C_DARK_BLUE, name="Arial", size=10)
        steps_load = [
            "1. Data  →  Queries & Connections  (panel kanan akan muncul)",
            f"2. Klik kanan '{query_name}'  →  Load To...",
            "3. Pilih 'Table'  →  pilih 'Existing worksheet'  →  klik cell A6 di sheet ini  →  OK",
            "4. Data akan muncul mulai baris 6",
            "5. Untuk refresh: Ctrl+Alt+F5 atau klik kanan tabel  →  Refresh",
        ]
        for si, step in enumerate(steps_load, 5):
            ws.cell(si, 1, step)
            ws.cell(si, 1).font = Font(color="374151", name="Arial", size=9)
            ws.row_dimensions[si].height = 16
        ws.row_dimensions[4].height = 18

    wb.save(out_path)
    print(f"[1/2] Workbook dasar disimpan: {out_path}")

    # ── Inject Power Query ────────────────────────────────────────────────────
    inject_power_query(out_path)
    print(f"[2/2] Power Query ter-inject. File siap: {out_path}")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    out = r"C:\Project\sj-monitor-firestore-query-v2.xlsx"
    build_workbook(out)
    size_kb = os.path.getsize(out) / 1024
    print(f"Ukuran file: {size_kb:.1f} KB")
    print("Selesai.")
